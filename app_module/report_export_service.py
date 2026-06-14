import os
import math
from pathlib import Path
from uuid import uuid4
from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List, Optional, Callable
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app_module.report_export_dtos import (
    ReportMetadata,
    SingleBacktestExportPayload,
    BatchBacktestExportPayload,
    RecommendationReplayExportPayload,
    CurrentRecommendationExportPayload,
)


class ReportExportService:
    """報告匯出服務 (Excel Exporter)
    提供規格化的報告匯出至 Excel 格式，具備專業樣式、凍結窗格、自動欄寬與原子寫入。
    """

    def __init__(self):
        # 樣式定義
        self.title_font = Font(name="Microsoft JhengHei", size=16, bold=True, color="1F4E79")
        self.section_font = Font(name="Microsoft JhengHei", size=12, bold=True, color="2C3E50")
        self.header_font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
        self.data_font = Font(name="Consolas", size=10)
        self.meta_label_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="555555")
        
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.meta_label_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
        self.integrity_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style="thin", color="BDC3C7"),
            right=Side(style="thin", color="BDC3C7"),
            top=Side(style="thin", color="BDC3C7"),
            bottom=Side(style="thin", color="BDC3C7"),
        )
        self.double_bottom_border = Border(
            bottom=Side(style="double", color="2C3E50"),
            top=Side(style="thin", color="BDC3C7"),
        )

    def export_single_backtest(self, target_path: Path, payload: SingleBacktestExportPayload) -> Path:
        """匯出單股回測報告"""
        return self._export(
            target_path,
            lambda wb: self._build_single_backtest_workbook(wb, payload)
        )

    def export_batch_backtest(self, target_path: Path, payload: BatchBacktestExportPayload) -> Path:
        """匯出批次回測報告"""
        return self._export(
            target_path,
            lambda wb: self._build_batch_backtest_workbook(wb, payload)
        )

    def export_recommendation_replay(self, target_path: Path, payload: RecommendationReplayExportPayload) -> Path:
        """匯出推薦回放報告"""
        return self._export(
            target_path,
            lambda wb: self._build_recommendation_replay_workbook(wb, payload)
        )

    def export_current_recommendation(self, target_path: Path, payload: CurrentRecommendationExportPayload) -> Path:
        """匯出當前推薦報告"""
        return self._export(
            target_path,
            lambda wb: self._build_current_recommendation_workbook(wb, payload)
        )

    def _export(self, target_path: Path, build_func: Callable[[openpyxl.Workbook], None]) -> Path:
        """通用匯出邏輯，包含路徑規格化與原子寫入"""
        target_path = self._normalize_target_path(target_path)
        
        # 確保父目錄存在
        if not target_path.parent.exists():
            raise FileNotFoundError(f"目標目錄不存在：{target_path.parent}")

        wb = openpyxl.Workbook()
        # 移除預設的 Sheet
        if wb.active:
            wb.remove(wb.active)

        build_func(wb)

        # 原子寫入
        tmp_path = target_path.with_name(f".{target_path.name}.{uuid4().hex}.tmp")
        try:
            wb.save(tmp_path)
            # 覆蓋或重新命名
            if os.path.exists(target_path):
                os.remove(target_path)
            os.rename(tmp_path, target_path)
        except Exception as e:
            if tmp_path.exists():
                tmp_path.unlink()
            raise e
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

        return target_path

    def _normalize_target_path(self, path: Path) -> Path:
        """規格化目標路徑，確保為 Path 對象且有 xlsx 副檔名"""
        p = Path(path)
        if p.suffix.lower() != ".xlsx":
            p = p.with_suffix(".xlsx")
        return p

    def _safe_excel_value(self, val: Any) -> Any:
        """將值轉換為 Excel 支援的安全型態"""
        if val is None:
            return "N/A"
        # 避免 list/dict/tuple 傳入 pd.isna 導致 truth value 歧義錯誤
        if isinstance(val, (list, dict, set, tuple)):
            if not val:
                return "N/A"
            return str(val)
        
        try:
            if pd.isna(val):
                return "N/A"
        except Exception:
            pass

        if isinstance(val, (int, float)):
            if math.isnan(val) or math.isinf(val):
                return "N/A"
            return val
        if isinstance(val, Decimal):
            # presentation boundary: 將 Decimal 轉為 float 供 Excel 計算
            f_val = float(val)
            if math.isnan(f_val) or math.isinf(f_val):
                return "N/A"
            return f_val
        if isinstance(val, datetime):
            # 移除時區，防止 openpyxl timezone support 報錯
            return val.replace(tzinfo=None)
        return str(val)


    def _write_key_value_section(self, ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, title: str, kv_dict: Dict[str, Any]) -> int:
        """寫入 Key-Value 區塊（例如參數或指標），回傳結束後的下一列列號"""
        ws.cell(row=start_row, column=1, value=title).font = self.section_font
        current_row = start_row + 1
        
        for k, v in kv_dict.items():
            cell_k = ws.cell(row=current_row, column=1, value=k)
            cell_k.font = self.meta_label_font
            cell_k.fill = self.meta_label_fill
            cell_k.border = self.thin_border
            
            cell_v = ws.cell(row=current_row, column=2, value=self._safe_excel_value(v))
            cell_v.font = self.data_font
            cell_v.border = self.thin_border
            
            # 格式化數值
            if isinstance(v, (int, float, Decimal)):
                if abs(float(v)) < 1.0 and float(v) != 0.0:
                    cell_v.number_format = "0.00%"
                else:
                    cell_v.number_format = "#,##0.00"
            
            current_row += 1
            
        return current_row + 1

    def _write_metadata_section(self, ws: openpyxl.worksheet.worksheet.Worksheet, metadata: ReportMetadata) -> int:
        """寫入元數據與資料完整性，回傳結束後下一列"""
        meta_dict = {
            "報告類型": metadata.report_type,
            "產生時間": metadata.generated_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "資料截止日期": metadata.data_as_of_date,
            "資料版本 (Data Version)": metadata.data_version,
            "策略代號 (Strategy ID)": metadata.strategy_id,
            "策略版本 (Strategy Version)": metadata.strategy_version,
            "市場 Regime": metadata.regime,
            "對比基準 (Benchmark)": metadata.benchmark,
            "交易執行假設": metadata.execution_assumption,
        }
        
        next_row = self._write_key_value_section(ws, 3, "報告追溯元數據 (Traceability)", meta_dict)
        
        # 檢查資料完整性
        missing = metadata.missing_fields()
        if missing:
            ws.cell(row=next_row, column=1, value="資料完整性警示 (Data Integrity)").font = self.section_font
            ws.cell(row=next_row + 1, column=1, value="缺失欄位：").font = self.meta_label_font
            ws.cell(row=next_row + 1, column=1).fill = self.integrity_fill
            ws.cell(row=next_row + 1, column=1).border = self.thin_border
            
            cell_missing = ws.cell(row=next_row + 1, column=2, value=", ".join(missing))
            cell_missing.font = Font(name="Consolas", size=10, bold=True, color="9C27B0")
            cell_missing.fill = self.integrity_fill
            cell_missing.border = self.thin_border
            next_row += 3
            
        return next_row

    def _write_dataframe(self, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame, freeze_pane: bool = True):
        """寫入 DataFrame 到 Sheet"""
        if df.empty:
            ws.cell(row=1, column=1, value="無相關數據資料").font = self.section_font
            return

        # 寫入 Header
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

        # 寫入 Data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._safe_excel_value(val))
                cell.font = self.data_font
                cell.border = self.thin_border
                
                # 自動判斷格式
                col_name = df.columns[col_idx - 1]
                if isinstance(val, (int, float, Decimal)):
                    # 百分比相關欄位
                    if "率" in col_name or "比" in col_name or "drawdown" in col_name or "weight" in col_name or "權重" in col_name or "貢獻" in col_name:
                        cell.number_format = "0.00%"
                    # 整數
                    elif "股" in col_name or "量" in col_name or "排名" in col_name:
                        cell.number_format = "#,##0"
                    # 金額/一般數值
                    else:
                        cell.number_format = "#,##0.00"

        # 凍結首列與啟用篩選
        if freeze_pane:
            ws.freeze_panes = "A2"
            
        # 自動縮放欄寬
        self._set_bounded_column_widths(ws)
        
        # 自動過濾器
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{df.shape[0]+1}"

    def _set_bounded_column_widths(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """根據內容設定欄寬，但限制最大欄寬為 60"""
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = cell.value
                if val is not None:
                    # 中文字元長度估算
                    val_str = str(val)
                    # 每包含一個中文，長度 + 1 模擬 double-width
                    ch_len = len(val_str) + sum(1 for c in val_str if '\u4e00' <= c <= '\u9fff')
                    max_len = max(max_len, ch_len)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

    # --- 報告 Workbook 建構子 ---

    def _build_single_backtest_workbook(self, wb: openpyxl.Workbook, payload: SingleBacktestExportPayload):
        """建立單股回測的 Sheet"""
        # Sheet 1: 摘要與設定
        ws_summary = wb.create_sheet(title="摘要與設定")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary.cell(row=1, column=1, value="單股回測研究報告").font = self.title_font
        
        next_row = self._write_metadata_section(ws_summary, payload.metadata)
        next_row = self._write_key_value_section(ws_summary, next_row, "回測執行參數 (Parameters)", payload.run_params)
        next_row = self._write_key_value_section(ws_summary, next_row, "策略績效指標 (Metrics)", payload.metrics)
        self._write_key_value_section(ws_summary, next_row, "回測驗證狀態 (Validation)", payload.validation)
        self._set_bounded_column_widths(ws_summary)

        # Sheet 2: 交易明細
        ws_trades = wb.create_sheet(title="交易明細")
        ws_trades.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_trades, payload.trades)

        # Sheet 3: 淨值與回撤
        ws_equity = wb.create_sheet(title="淨值與回撤")
        ws_equity.views.sheetView[0].showGridLines = True
        
        eq_df = payload.equity_curve.copy()
        if not eq_df.empty:
            # 如果沒有 drawdown，自動計算
            if "drawdown" not in eq_df.columns:
                if "equity" in eq_df.columns:
                    eq_df["drawdown"] = self._build_drawdown_series(eq_df["equity"])
                else:
                    eq_df["drawdown"] = 0
            
            # 保證 drawdown 欄位存在
            if "equity" in eq_df.columns and "drawdown" in eq_df.columns:
                # 重新排列欄位，使順序為 日期, equity, drawdown
                cols = ["日期", "equity", "drawdown"]
                # 把其他可能存在的欄位也塞在後面
                cols.extend([c for c in eq_df.columns if c not in cols])
                eq_df = eq_df[cols]
                
        self._write_dataframe(ws_equity, eq_df)

    def _build_drawdown_series(self, equity_series: pd.Series) -> pd.Series:
        """輔助產生 drawdown 序列"""
        # 將 Decimal 轉為 float
        vals = equity_series.apply(lambda x: float(x) if isinstance(x, Decimal) else x)
        peaks = vals.cummax()
        # 避免除以 0
        drawdowns = (vals - peaks) / peaks.replace(0, 1.0)
        return drawdowns

    def _build_batch_backtest_workbook(self, wb: openpyxl.Workbook, payload: BatchBacktestExportPayload):
        """建立批次回測的 Sheet"""
        # Sheet 1: 批次總覽
        ws_summary = wb.create_sheet(title="批次總覽")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary.cell(row=1, column=1, value="批次操作回測總覽").font = self.title_font
        
        next_row = self._write_metadata_section(ws_summary, payload.metadata)
        
        # 寫入整體統計
        ws_summary.cell(row=next_row, column=1, value="整體統計資訊").font = self.section_font
        cell_stats = ws_summary.cell(row=next_row + 1, column=1, value=payload.overall_stats)
        cell_stats.font = self.data_font
        cell_stats.alignment = Alignment(wrap_text=True)
        ws_summary.row_dimensions[next_row + 1].height = 60
        self._set_bounded_column_widths(ws_summary)

        # Sheet 2: 排行榜
        ws_leader = wb.create_sheet(title="排行榜")
        ws_leader.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_leader, payload.leaderboard)

        # Sheet 3: 失敗與警告
        ws_fails = wb.create_sheet(title="失敗與警告")
        ws_fails.views.sheetView[0].showGridLines = True
        
        fail_df = pd.DataFrame()
        if not payload.leaderboard.empty:
            # 篩選 status != 'SUCCESS' 或是 '狀態' 欄位有 FAILED/ERROR 的資料
            status_col = None
            for col in payload.leaderboard.columns:
                if "狀態" in col or "status" in col.lower():
                    status_col = col
                    break
            
            if status_col:
                fail_df = payload.leaderboard[payload.leaderboard[status_col].astype(str).str.upper() != "SUCCESS"]
                
        self._write_dataframe(ws_fails, fail_df)

    def _build_recommendation_replay_workbook(self, wb: openpyxl.Workbook, payload: RecommendationReplayExportPayload):
        """建立推薦回放報告的 Sheet"""
        # Sheet 1: 回放摘要與設定
        ws_summary = wb.create_sheet(title="回放摘要與設定")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary.cell(row=1, column=1, value="推薦回放研究報告").font = self.title_font
        
        next_row = self._write_metadata_section(ws_summary, payload.metadata)
        next_row = self._write_key_value_section(ws_summary, next_row, "回放執行參數 (Parameters)", payload.run_params)
        next_row = self._write_key_value_section(ws_summary, next_row, "回放績效指標 (Metrics)", payload.summary)
        
        # 寫入診斷與改進提示
        if payload.diagnostics:
            ws_summary.cell(row=next_row, column=1, value="回放診斷資訊 (Diagnostics)").font = self.section_font
            next_row += 1
            for diag in payload.diagnostics:
                ws_summary.cell(row=next_row, column=1, value="• " + diag).font = self.data_font
                next_row += 1
            next_row += 1
            
        if payload.improvement_hints:
            ws_summary.cell(row=next_row, column=1, value="回放改進建議 (Improvement Hints)").font = self.section_font
            next_row += 1
            for hint in payload.improvement_hints:
                ws_summary.cell(row=next_row, column=1, value="• " + hint).font = self.data_font
                next_row += 1
                
        self._set_bounded_column_widths(ws_summary)

        # Sheet 2: 期間持倉
        ws_holdings = wb.create_sheet(title="期間持倉")
        ws_holdings.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_holdings, payload.period_holdings)

        # Sheet 3: 個股貢獻
        ws_contrib = wb.create_sheet(title="個股貢獻")
        ws_contrib.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_contrib, payload.stock_contribution)

        # Sheet 4: 交易紀錄
        ws_trades = wb.create_sheet(title="交易紀錄")
        ws_trades.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_trades, payload.trades)

        # Sheet 5: 淨值與回撤
        ws_equity = wb.create_sheet(title="淨值與回撤")
        ws_equity.views.sheetView[0].showGridLines = True
        
        eq_df = payload.equity_curve.copy()
        if not eq_df.empty:
            if "drawdown" not in eq_df.columns:
                if "equity" in eq_df.columns:
                    eq_df["drawdown"] = self._build_drawdown_series(eq_df["equity"])
                else:
                    eq_df["drawdown"] = 0
            
            if "equity" in eq_df.columns and "drawdown" in eq_df.columns:
                cols = ["日期", "equity", "drawdown"]
                cols.extend([c for c in eq_df.columns if c not in cols])
                eq_df = eq_df[cols]
                
        self._write_dataframe(ws_equity, eq_df)

    def _build_current_recommendation_workbook(self, wb: openpyxl.Workbook, payload: CurrentRecommendationExportPayload):
        """建立當前推薦的 Sheet"""
        # Sheet 1: 推薦總覽與配置
        ws_summary = wb.create_sheet(title="推薦總覽與配置")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary.cell(row=1, column=1, value="今日推薦與配置報告").font = self.title_font
        
        next_row = self._write_metadata_section(ws_summary, payload.metadata)
        next_row = self._write_key_value_section(ws_summary, next_row, "執行參數 (Parameters)", payload.run_params)
        
        if payload.regime_snapshot:
            self._write_key_value_section(ws_summary, next_row, "今日市場狀態快照 (Regime Snapshot)", payload.regime_snapshot)
            
        self._set_bounded_column_widths(ws_summary)

        # Sheet 2: 推薦股票名單
        ws_recs = wb.create_sheet(title="推薦股票名單")
        ws_recs.views.sheetView[0].showGridLines = True
        self._write_dataframe(ws_recs, payload.recommendations)
