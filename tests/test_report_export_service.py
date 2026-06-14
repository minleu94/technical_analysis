import pytest
import os
from pathlib import Path
import pandas as pd
import openpyxl
from decimal import Decimal
from datetime import datetime

from app_module.report_export_dtos import (
    ReportMetadata,
    SingleBacktestExportPayload,
    BatchBacktestExportPayload,
    RecommendationReplayExportPayload,
    CurrentRecommendationExportPayload,
)
from app_module.report_export_service import ReportExportService


@pytest.fixture
def exporter():
    return ReportExportService()


@pytest.fixture
def sample_metadata():
    return ReportMetadata(
        report_type="single_backtest",
        generated_at="2026-06-14T12:00:00",
        data_as_of_date="2026-06-12",
        data_version="sha256:test1234",
        strategy_id="test_strat",
        strategy_version="1.0",
        regime="Trend",
        benchmark="TAIEX",
        execution_assumption="next_open",
    )


def test_export_single_backtest(tmp_path, exporter, sample_metadata):
    target = tmp_path / "single_report.xlsx"
    trades = pd.DataFrame([
        {"交易日期": "2026-06-01", "證券代號": "2330", "證券名稱": "台積電", "交易動作": "BUY", "成交價格": Decimal("800.5"), "成交股數": 1000}
    ])
    equity_curve = pd.DataFrame([
        {"日期": "2026-06-01", "equity": Decimal("100000")}
    ])
    
    payload = SingleBacktestExportPayload(
        metadata=sample_metadata,
        run_params={"fee_bps": 14.25, "slippage_bps": 5.0},
        metrics={"total_return": Decimal("0.125"), "max_drawdown": Decimal("-0.05")},
        validation={"status": "PASS", "messages": []},
        trades=trades,
        equity_curve=equity_curve,
    )
    
    output_path = exporter.export_single_backtest(target, payload)
    assert output_path.exists()
    
    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["摘要與設定", "交易明細", "淨值與回撤"]
    
    # 驗證 A1 標題
    assert wb["摘要與設定"]["A1"].value == "單股回測研究報告"
    # 驗證 metadata 有被寫入
    sheet_meta = wb["摘要與設定"]
    meta_text = ""
    for r in range(1, 30):
        for c in (1, 2):
            val = sheet_meta.cell(row=r, column=c).value
            if val:
                meta_text += str(val) + " "
    assert "Trend" in meta_text

    
    # 驗證淨值與回撤自動產生 drawdown 欄位 (因為 source 只有 equity)
    assert wb["淨值與回撤"]["C1"].value == "drawdown"


def test_export_single_backtest_missing_metadata(tmp_path, exporter):
    target = tmp_path / "single_report_missing.xlsx"
    # 故意留空一些欄位以測試缺失欄位
    metadata = ReportMetadata(
        report_type="single_backtest",
        generated_at="2026-06-14T12:00:00",
    )
    payload = SingleBacktestExportPayload(
        metadata=metadata,
        run_params={},
        metrics={},
        validation={},
        trades=pd.DataFrame(),
        equity_curve=pd.DataFrame([{"日期": "2026-06-01", "equity": 100, "drawdown": 0}]),
    )
    
    output_path = exporter.export_single_backtest(target, payload)
    wb = openpyxl.load_workbook(output_path)
    # 驗證資料完整性區塊
    sheet = wb["摘要與設定"]
    text = ""
    for r in range(1, 30):
        for c in range(1, 5):
            val = sheet.cell(row=r, column=c).value
            if val:
                text += str(val) + " "
    assert "資料完整性" in text
    assert "缺失欄位" in text
    assert "data_as_of_date" in text


def test_export_batch_backtest(tmp_path, exporter, sample_metadata):
    target = tmp_path / "batch_report.xlsx"
    leaderboard = pd.DataFrame([
        {"排名": 1, "證券代號": "2330", "總報酬率": 0.25, "最大回撤": -0.04, "狀態": "SUCCESS"},
        {"排名": 2, "證券代號": "2454", "總報酬率": -0.05, "最大回撤": -0.15, "狀態": "FAILED"}
    ])
    
    payload = BatchBacktestExportPayload(
        metadata=sample_metadata,
        leaderboard=leaderboard,
        overall_stats={
            "total_stocks": 2,
            "successful_stocks": 1,
            "failed_stocks": 1,
        },
    )
    
    output_path = exporter.export_batch_backtest(target, payload)
    assert output_path.exists()
    
    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["批次總覽", "排行榜", "失敗與警告"]
    
    assert wb["批次總覽"]["A1"].value == "批次操作回測總覽"
    # 排行榜應該只包含成功的或全部？
    # 按照 test 契約，「排行榜」通常寫所有資料，但「失敗與警告」則是只寫失敗或警告的資料
    # 我們讓「排行榜」包含所有 leaderboard 資料，而「失敗與警告」過濾出 status != 'SUCCESS' 或是 FAILED 的行
    assert wb["排行榜"]["B2"].value == "2330"
    assert wb["失敗與警告"]["B2"].value == "2454"


def test_export_recommendation_replay(tmp_path, exporter, sample_metadata):
    target = tmp_path / "replay_report.xlsx"
    payload = RecommendationReplayExportPayload(
        metadata=sample_metadata,
        run_params={"initial_capital": 1000000},
        summary={"total_return": 0.15, "win_rate": 0.6},
        period_holdings=pd.DataFrame([{"日期": "2026-06-01", "證券代號": "2330", "權重": 0.5}]),
        stock_contribution=pd.DataFrame([{"證券代號": "2330", "貢獻度": 0.08}]),
        trades=pd.DataFrame([{"交易日期": "2026-06-01", "動作": "BUY"}]),
        equity_curve=pd.DataFrame([{"日期": "2026-06-01", "equity": 1000000, "drawdown": 0}]),
        diagnostics=["診斷訊息1"],
        improvement_hints=["改進建議1"]
    )
    
    output_path = exporter.export_recommendation_replay(target, payload)
    assert output_path.exists()
    
    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["回放摘要與設定", "期間持倉", "個股貢獻", "交易紀錄", "淨值與回撤"]
    assert wb["回放摘要與設定"]["A1"].value == "推薦回放研究報告"


def test_export_current_recommendation(tmp_path, exporter, sample_metadata):
    target = tmp_path / "current_recommendation_report.xlsx"
    recs = pd.DataFrame([
        {"證券代號": "2330", "證券名稱": "台積電", "評分": 85.5, "推薦理由": "技術面強勢，突破均線，成交量顯著放大"}
    ])
    payload = CurrentRecommendationExportPayload(
        metadata=sample_metadata,
        run_params={"min_score": 80},
        recommendations=recs,
        regime_snapshot={"regime": "Trend", "confidence": 0.9}
    )
    
    output_path = exporter.export_current_recommendation(target, payload)
    assert output_path.exists()
    
    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["推薦總覽與配置", "推薦股票名單"]
    assert wb["推薦總覽與配置"]["A1"].value == "今日推薦與配置報告"


def test_atomic_write_failure_cleanup(tmp_path, exporter, sample_metadata):
    target = tmp_path / "fail_report.xlsx"
    
    payload = SingleBacktestExportPayload(
        metadata=sample_metadata,
        run_params={},
        metrics={},
        validation={},
        trades=pd.DataFrame([{"A": 1}]),
        equity_curve=pd.DataFrame([{"B": 1}]),
    )
    
    # 使用 mock patch 讓 _write_dataframe 拋出 Exception
    import unittest.mock as mock
    with mock.patch.object(exporter, "_write_dataframe", side_effect=ValueError("forced error")):
        with pytest.raises(ValueError):
            exporter.export_single_backtest(target, payload)
        
    # 驗證不會留下目標檔，也不會留下任何 .tmp 暫存檔
    assert not target.exists()
    tmp_files = list(tmp_path.glob("*.tmp"))
    assert len(tmp_files) == 0


def test_atomic_replace_failure_preserves_existing_report(
    tmp_path, exporter, sample_metadata, monkeypatch
):
    target = tmp_path / "existing_report.xlsx"
    original = b"existing report"
    target.write_bytes(original)
    payload = SingleBacktestExportPayload(
        metadata=sample_metadata,
        run_params={},
        metrics={},
        validation={},
        trades=pd.DataFrame(),
        equity_curve=pd.DataFrame(),
    )

    def fail_replace(source, destination):
        raise PermissionError("target is locked")

    monkeypatch.setattr(os, "replace", fail_replace)

    with pytest.raises(PermissionError, match="target is locked"):
        exporter.export_single_backtest(target, payload)

    assert target.read_bytes() == original
    assert list(tmp_path.glob("*.tmp")) == []


@pytest.mark.parametrize(
    "equity_curve",
    [
        pd.DataFrame([{"date": "2026-06-01", "equity": 100000}]),
        pd.DataFrame(
            {"equity": [100000]},
            index=pd.Index(["2026-06-01"], name="date"),
        ),
        pd.DataFrame(
            {"equity": [100000]},
            index=pd.Index(["2026-06-01"]),
        ),
    ],
)
def test_single_export_normalizes_real_equity_curve_date_shapes(
    tmp_path, exporter, sample_metadata, equity_curve
):
    target = tmp_path / "real_equity_shape.xlsx"
    payload = SingleBacktestExportPayload(
        metadata=sample_metadata,
        run_params={},
        metrics={},
        validation={},
        trades=pd.DataFrame(),
        equity_curve=equity_curve,
    )

    exporter.export_single_backtest(target, payload)

    workbook = openpyxl.load_workbook(target)
    sheet = workbook["淨值與回撤"]
    assert sheet["A1"].value == "日期"
    assert sheet["A2"].value == "2026-06-01"
    assert sheet["B1"].value == "equity"


def test_batch_export_failure_sheet_supports_real_leaderboard_shape(
    tmp_path, exporter, sample_metadata
):
    target = tmp_path / "batch_real_shape.xlsx"
    payload = BatchBacktestExportPayload(
        metadata=sample_metadata,
        leaderboard=pd.DataFrame(
            [
                {"股票代號": "2330", "CAGR%": 25.0, "失敗原因": ""},
                {"股票代號": "2454", "CAGR%": None, "失敗原因": "資料不足"},
            ]
        ),
        overall_stats={"total_stocks": 2, "successful_stocks": 1},
    )

    exporter.export_batch_backtest(target, payload)

    workbook = openpyxl.load_workbook(target)
    failure_sheet = workbook["失敗與警告"]
    assert failure_sheet["A2"].value == "2454"
    assert failure_sheet["C2"].value == "資料不足"


def test_safe_excel_value_conversion(tmp_path, exporter, sample_metadata):
    target = tmp_path / "conversion_report.xlsx"
    
    # 測試包含 Inf, NaN, None, tz-datetime, Decimal 的寫入
    trades = pd.DataFrame([
        {
            "證券代號": "2330",
            "金額": Decimal("100.50"),
            "無效值": float("nan"),
            "極限值": float("inf"),
            "日期": datetime.now(),
        }
    ])
    
    payload = SingleBacktestExportPayload(
        metadata=sample_metadata,
        run_params={},
        metrics={},
        validation={},
        trades=trades,
        equity_curve=pd.DataFrame([{"日期": "2026-06-01", "equity": 100, "drawdown": 0}]),
    )
    
    output_path = exporter.export_single_backtest(target, payload)
    assert output_path.exists()
    
    wb = openpyxl.load_workbook(output_path)
    sheet = wb["交易明細"]
    # nan 應該轉換成 "N/A" 或是空值？
    # 根據 plan，缺值或無效值輸出 "N/A"
    assert sheet.cell(row=2, column=2).value == 100.50  # Decimal 轉成 float
    assert sheet.cell(row=2, column=3).value == "N/A"   # NaN 轉成 "N/A"
    assert sheet.cell(row=2, column=4).value == "N/A"   # Inf 轉成 "N/A"
